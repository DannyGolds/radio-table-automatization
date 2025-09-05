import openpyxl
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
import xlrd



class Comparing:
    def __init__(self, file1, file2, cols_to_compare, output_path):
        self.file1 = file1
        self.file2 = file2
        self.output_path = output_path
        self.cols_to_compare = [self.column_letter_to_index(col) for col in cols_to_compare]
        self.__titles = ['Технические характеристики', 'Существующее оборудование', 'Проектируемое оборудование', 'БС']
        self.__key_of_verticalize = "№ антенны"
        self.__sizes_of_cols = [6.79, 10.95, 4.43, 4.43, 4.43, 4.43, 6.37, 6.37, 14.37, 4.43, 4.43, 4.43, 4.43, 5.57, 6.79, 4.43, 5.71, 6.29, 5.43, 6.71, 6, 7, 7]
        self.__column_mapping = {}
        self.__cells_to_fill = []


    def open_workbook(self, input_path):
        xl_wb = xlrd.open_workbook(input_path)
        xl_sheet = xl_wb.sheet_by_index(0)
        
        op_wb = openpyxl.Workbook()
        op_sheet = op_wb.active
        
        for row in range(xl_sheet.nrows):
            for col in range(xl_sheet.ncols):
                op_sheet.cell(row=row+1, column=col+1).value = xl_sheet.cell_value(row, col)
        
        return op_wb
    

    def column_letter_to_index(self, letter):
        letter = letter.upper()
        index = 0

        for char in letter:
            if 'A' <= char <= 'Z':
                index = index * 26 + (ord(char) - ord('A') + 1)
            else:
                raise ValueError(f"Invalid column letter: {letter}")
            
        return index
    
    def copy_data(self, source_sheet, target_sheet):
        for row in range(1, source_sheet.max_row + 1):
            for col in range(1, source_sheet.max_column + 1):
                value = source_sheet.cell(row=row, column=col).value
                if isinstance(value, float):
                    value = round(value, 2)
                target_sheet.cell(row=row, column=col, value=value)

        
        offset = 0
        for col in sorted(self.cols_to_compare):
            ## Вставляем новый столбец после оригинального
            target_sheet.insert_cols(col + 1 + offset)
            self.__column_mapping[col] = col + 1 + offset
            offset +=1

        ## Возвращаем количество столбцов, добавленных для сравнения(чтобы откорректировать количество столбцов в итоговом файле)
                

    """Функция сравнивания двух файлов Excel и выделения различий"""
    def compare(self):
        try:
            # Открываем файлы с помощью xlrd и openpyxl
            self.__wb1 = self.open_workbook(self.file1)
            self.__wb2 = self.open_workbook(self.file2)
            self.__sheet1 = self.__wb1.active
            self.__sheet2 = self.__wb2.active
            self.__result_wb = openpyxl.Workbook()
            self.__result_sheet = self.__result_wb.active
            self.__result_sheet.title = "Результат сравнения"
            self.copy_data(self.__sheet1, self.__result_sheet)
            for row in range(1, min(self.__sheet1.max_row, self.__sheet2.max_row) + 1):
                for ofor, new_col in self.__column_mapping.items():
                    ## ofor - original file original result in original file position, not in result file
                    orig_col = new_col - 1
                    val1 = self.__sheet1.cell(row=row, column=ofor).value
                    val2 = self.__sheet2.cell(row=row, column=ofor).value
                    
                    
                    if val1 != val2:
                        self.__result_sheet.cell(row=row, column=new_col, value=round(val2, 2))
                        self.__cells_to_fill.append((row, new_col))
                    else:
                        temp = self.__result_sheet.cell(row=row, column=orig_col).value
                        self.__result_sheet.merge_cells(start_row=row, end_row=row, start_column=orig_col, end_column=new_col)
                        self.__result_sheet.cell(row=row, column=orig_col, value=temp)

            return True, "Сравнение завершено успешно"

        except Exception as e:
            return False, f"Ошибка при открытии файлов: {str(e)}"
        
    
    """Здесь будет логика стилизации ячеек"""
    def stylization(self):

        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        red_font = Font(color="FF0000", size=11.0, name="Times New Roman")  # Красный цвет текста
        font = Font(color="000000", size=11.0, name="Times New Roman")
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        vertical_alignment = Alignment(text_rotation=90, vertical='center', horizontal="center", wrap_text=True)
        was_the_first_title = False
        # Устанавливаем узкие границы и формат A4
        # Устанавливаем поля как в Excel "Узкие":
        # Сверху/Снизу: 1.91 см = 0.75 дюйма
        # Слева/Справа: 0.64 см = 0.25 дюйма
        # Верхний/Нижний колонтитул: 0.76 см = 0.3 дюйма
        self.__result_sheet.page_margins = PageMargins(
            left=0.25,   # 0.25 inch = 0.64 cm
            right=0.25,
            top=0.75,    # 0.75 inch = 1.91 cm
            bottom=0.75,
            header=0.3,  # 0.3 inch = 0.76 cm
            footer=0.3
        )
        self.__result_sheet.page_setup.orientation = 'landscape'
        self.__result_sheet.page_setup.paperSize = 9  # 9 = A4
        

        for row in range(1, self.__result_sheet.max_row + 1):
            self.__result_sheet.row_dimensions[row].height = 30  # Высота строки

            if self.__result_sheet.cell(row=row, column=1).value.strip().startswith("БС"):
                self.__result_sheet.merge_cells(start_row=row, end_row=row, start_column=1, end_column=self.__result_sheet.max_column)
                self.__result_sheet.cell(row=row, column=1).alignment = left_align
            
            if self.__titles[0] in self.__result_sheet.cell(row=row, column=1).value.strip():
                if not was_the_first_title:
                    was_the_first_title = True
                    self.__result_sheet.merge_cells(start_row=row, end_row=row, start_column=1, end_column = self.__result_sheet.max_column)
                    self.__result_sheet.cell(row=row, column=1).alignment = center_alignment
                    self.__result_sheet.cell(row=row, column=1).border = thin_border
                    self.__result_sheet.cell(row=row, column=1).font = font
                    self.__result_sheet.cell(row=row, column=self.__result_sheet.max_column).border = thin_border
                    continue
                self.__result_sheet.row_breaks.append(Break(id=row - 1))

            if self.__result_sheet.cell(row=row, column=1).value \
            and (str(self.__result_sheet.cell(row=row, column=1).value).lower().startswith(self.__titles[0].lower())\
            or str(self.__result_sheet.cell(row=row, column=1).value).lower().startswith(self.__titles[1].lower())\
            or str(self.__result_sheet.cell(row=row, column=1).value).lower().startswith(self.__titles[2].lower())):
                self.__result_sheet.merge_cells(start_row=row, end_row=row, start_column=1, end_column = self.__result_sheet.max_column)
            
            if self.__key_of_verticalize.lower() in self.__result_sheet.cell(row=row, column=1).value.strip().lower():
                for col in range (1, self.__result_sheet.max_column + 1):
                    if self.__result_sheet.cell(row=row, column=col).value and str(self.__result_sheet.cell(row=row, column=col).value).strip().lower() != 'Тип антенны'.lower():
                        self.__result_sheet.cell(row=row, column=col).alignment = vertical_alignment
                    else:
                        self.__result_sheet.cell(row=row, column=col).alignment = center_alignment

            if self.__result_sheet.cell(row=row, column=1).value and self.__key_of_verticalize.lower() == str(self.__result_sheet.cell(row=row, column=1).value).strip().lower():
                self.__result_sheet.row_dimensions[row].height = 132  # Высота строки
                for col in range (1, self.__result_sheet.max_column + 1):
                    if self.__result_sheet.cell(row=row, column=col).value and str(self.__result_sheet.cell(row=row, column=col).value).strip().lower() != 'Тип антенны'.lower():
                        self.__result_sheet.cell(row=row, column=col).alignment = vertical_alignment
            for col in range(1, self.__result_sheet.max_column + 1):
                self.__result_sheet.cell(row=row, column=col).border = thin_border
                if self.__result_sheet.cell(row=row, column=1).value and self.__key_of_verticalize.lower() != str(self.__result_sheet.cell(row=row, column=1).value).strip().lower():
                    self.__result_sheet.cell(row=row, column=col).alignment = center_alignment
                if (row, col) not in self.__cells_to_fill: 
                    self.__result_sheet.cell(row=row, column=col).font = font
                if (row, col) in self.__cells_to_fill:
                    self.__result_sheet.cell(row=row, column=col).font = red_font



        for col in range(0, self.__result_sheet.max_column + 1):
            self.__result_sheet.column_dimensions[get_column_letter(col + 1)].width = (self.__sizes_of_cols[col]* 7 + 5) / 7

            

    
    def save(self):
        self.stylization()
        try:
            self.__result_wb.save(self.output_path)
            return True, f"Файл успешно сохранен: {self.output_path}"
        except Exception as e:
            return False, f"Ошибка при сохранении файла: {str(e)}"
  
